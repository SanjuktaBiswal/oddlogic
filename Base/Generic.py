import json
import pickle

from bs4 import BeautifulSoup
from selenium.webdriver import ActionChains
from importlib import reload
from selenium.webdriver.common.by import By
from traceback import print_stack
from datetime import datetime
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import NamedStyle, Alignment, PatternFill,Font
import logging
import logging.config
import time
import os
import pandas as pd
import allure
from Utilities.configReader import ConfigReader


class Generic():

    def __init__(self, config,config_data,excel,testdata,driver,scriptName,className):
        self.driver = driver
        self.parentHandle=driver.current_window_handle
        self.testdata =testdata
        self.config = config
        self.excel=excel
        self.idcounter = 1
        self.li_idcounter = []
        self.config_data=config_data
        logDirectory = ".\\Logs\\"
        self.currentDirectory = os.path.dirname(__file__)
        logFolder = os.path.join(self.currentDirectory, logDirectory)
        if not os.path.isdir(logFolder):
            os.makedirs(logFolder)
        self.config_data.sLogFilePath=str(logFolder)
        self.config_data.sScriptFolderName=scriptName
        self.scriptName=scriptName
        self.config_data.projectpath=os.getcwd()#os.path.join(os.getcwd(),os.pardir)
        
        basic_config = ConfigReader(os.path.join(self.config_data.projectpath,"Config Files","Basic_Setup.ini"))
        self.log_type = basic_config.get_data_from_a_section("basic info", "log_type")
        self.override_screenshot=basic_config.get_data_from_a_section("basic info", "override_screenshot")
        self.screenshot_for_pass = basic_config.get_data_from_a_section("basic info", "screenshot_for_pass")
        self.screenshot_for_fail = basic_config.get_data_from_a_section("basic info", "screenshot_for_fail")
        self.className=className
        self.log=None
        if "txt" in self.log_type.strip().lower() or "log" in self.log_type.strip().lower():
            sDate = self.func_GetDateInDDMMYYYYFormat();
            self.logPath_log = os.path.join(self.config_data.sLogFilePath, self.config_data.sScriptFolderName, sDate,
                                        self.config_data.sLogFileName_log)

            self.logDirectory_log = os.path.join(self.config_data.sLogFilePath, self.config_data.sScriptFolderName, sDate)


            if not os.path.exists(self.logDirectory_log):
                os.makedirs(self.logDirectory_log)


            logging.config.fileConfig(os.path.join(self.config_data.projectpath,"Config Files","Basic_Setup.ini"),disable_existing_loggers=True,  defaults={'logfilename': str(self.logPath_log),'simple':'simpleFormatter'})
            #To get a new logger instance
            self.log = logging.getLogger(className)

        #-------------------------------------------
        bd = Side(border_style='thin')
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="top")
        center_alignment = Alignment(horizontal="center", vertical="center")

        greenfont = Font(color="FF00FF00")
        redfont = Font(color="FF0000")
        bluefont = Font(color="FF0000FF")
        orangefont = Font(color="FFFF6600")

        self.con_log_border = NamedStyle(name="con_log_border")

        self.con_log_border.border = border
        self.con_log_border.alignment = center_alignment


        self.log_border = NamedStyle(name="log_border")
        self.log_border.border = border
        self.log_border.alignment = left_alignment

        self.log_pass = NamedStyle(name="log_pass")
        self.log_pass.border = border
        self.log_pass.alignment = left_alignment
        self.log_pass.font = greenfont

        self.log_fail = NamedStyle(name="log_fail")
        self.log_fail.border = border
        self.log_fail.alignment = left_alignment
        self.log_fail.font = redfont
        self.log_fail.font.bold = True
        self.log_fail.font.size = 12


        self.log_info = NamedStyle(name="log_info")
        self.log_info.border = border
        self.log_info.alignment = left_alignment
        self.log_info.font = bluefont

        self.log_comment = NamedStyle(name="log_comment")
        self.log_comment.border = border
        self.log_comment.alignment = left_alignment
        self.log_comment.font = orangefont

        self.log_set = NamedStyle(name="log_set")
        self.log_set.border = border
        self.log_set.alignment = center_alignment
        self.log_set.fill = PatternFill(start_color="FFFFFF00", fill_type="solid")

        self.Logheader = NamedStyle(name="Logheader")

        self.Logheader.border = border
        self.Logheader.alignment = center_alignment
        self.Logheader.fill = PatternFill(start_color="FF7F3F", fill_type="solid")
        self.Logheader.alignment.wrap_text = True
        if self.config_data.sStartTime == "":
            self.config_data.sStartTime = self.func_GetTime();
        self.config_file = ConfigReader(config)

    def changeFormat(self):
        logging.config.fileConfig(os.path.join(self.config_data.projectpath,"Config Files","Basic_Setup.ini"),disable_existing_loggers=False,
                                  defaults={'logfilename': str(self.logPath_log), 'simple': 'simpleFormatter'})
        # To get a new logger instance

    def switchToNewWindow(self):
        # Find all handles, there should two handles after clicking open window button
        handles = self.driver.window_handles

        # Switch to window and search course
        for handle in handles:

            if handle not in self.parentHandle:
                self.driver.switch_to.window(handle)


                break
    def  markTheEndOfTextlog(self,msg):
        logging.shutdown()
        reload(logging)
        logging.config.fileConfig(os.path.join(self.config_data.projectpath,"Config Files","Basic_Setup.ini"), disable_existing_loggers=True,
                                  defaults={'logfilename': str(self.logPath_log),
                                            'simple': 'resultFormatter'})
        self.log = logging.getLogger(self.className)
        self.log.info(msg)
        logging.shutdown()
        reload(logging)
        logging.config.fileConfig(os.path.join(self.config_data.projectpath,"Config Files","Basic_Setup.ini"), disable_existing_loggers=True,
                                  defaults={'logfilename': str(self.logPath_log), 'simple': 'simpleFormatter'})
        # To get a new logger instance

        self.log = logging.getLogger(self.className)
    def slide(self,locator,x_coord, y_coord):

        slider=self.driver.find_element(By.XPATH, locator)
        loc=slider.location
        size=slider.size
        w,h=size['width'],size['height']
        ActionChains(self.driver).drag_and_drop_by_offset(slider,x_coord,y_coord).perform()
    def slide_half_x(self,locator):

        slider=self.driver.find_element(By.XPATH, locator)
        loc=slider.location
        size=slider.size
        w,h=size['width'],size['height']
        ActionChains(self.driver).drag_and_drop_by_offset(slider,w/2,0).perform()
    def slide_half_y(self,locator):

        slider=self.driver.find_element(By.XPATH, locator)
        loc=slider.location
        size=slider.size
        w,h=size['width'],size['height']
        ActionChains(self.driver).drag_and_drop_by_offset(slider,0,h/2).perform()
    def slide_full_x(self,locator):

        slider=self.driver.find_element(By.XPATH, locator)
        loc=slider.location
        size=slider.size
        w,h=size['width'],size['height']
        ActionChains(self.driver).drag_and_drop_by_offset(slider,w,0).perform()
    def slide_full_y(self,locator):

        slider=self.driver.find_element(By.XPATH, locator)
        loc=slider.location
        size=slider.size
        w,h=size['width'],size['height']
        ActionChains(self.driver).drag_and_drop_by_offset(slider,0,h).perform()
    def drag(self,dragable,dropable):

        dragable=self.driver.find_element(By.XPATH, dropable)
        dropable = self.driver.find_element(By.XPATH, dropable)
        ActionChains(self.driver).drag_and_drop(dragable,dropable).perform()
    def rightClick(self,field):
        title = self.driver.title.strip()

        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)

        locator_details = []
        locators = self.dict_screen_details[field]

        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)
        actionchains=ActionChains(self.driver)
        actionchains.context_click(self.driver.find_element(By.XPATH, locator_details[0])).perform()

    def scrollup(self,x_coord, y_coord):

        #x_coord is the horizontal pixel value that you want to scroll by.
        #y_coord is the vertical pixel value that you want to scroll by.

        self.driver.execute_script("window.scrollBy(-"+str(x_coord)+",-"+str(y_coord)+");")
        time.sleep(2)

    def scrolldown(self, x_coord, y_coord):
            # x_coord is the horizontal pixel value that you want to scroll by.
            # y_coord is the vertical pixel value that you want to scroll by.

            self.driver.execute_script("window.scrollBy(" + str(x_coord) + "," + str(y_coord) + ");")
            time.sleep(2)
        #self.driver.execute_script("window.scrollBy(0,"+str(y_coord)+");")
    def switchToIFrameUsingID(self,field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)

        self.func_InformationMessageToPrintManualSteps("Switch To IFrame");
        # Switch to frame using Id

        #self.driver.switch_to.frame("courses-iframe")
        self.driver.switch_to.frame("'"+locator_details[0]+"'")
    def swithBackToParentWindowFromIFrame(self):
        # Switch back to the parent frame
        self.driver.switch_to.default_content()
    def closeOpenedWindow(self):
        self.driver.close()
    def switchToParentWindow(self):
        self.driver.switch_to.window(self.parentHandle)
    def getTCName(self,iTCCounter) :

        sTemp = self.TCName[iTCCounter];

        return sTemp;
    def setLogMsgWithoutScreenCapture( self,sMsg) :
        try:
            self.config_data.sLogMsg.append(sMsg);
            self.config_data.sTCStatus.append("INFORMATION");
            self.config_data.arSCreenNames.append("");
            self.config_data.iTotalScreenCounter.append(0);
            self.config_data.sStepExecutionDate.append(self.func_GetDate());
            self.config_data.TCWithStats.put[self.config_data.sTestcaseId] = "Pass";

        except:
            pass
    def printSetNo(self):

        stTCID = self.config_data.sTestcaseId.split("_");
        TCID=stTCID[len(stTCID)-1]
        self.setLogMsgWithoutScreenCapture("Set: " + TCID.replace("S", "")
                                       + "\n" + "Test Case ID: "
                                       + self.config_data.sTestcaseId);

    def getByType(self, locatorType):
        locatorType = locatorType.lower()
        if locatorType == "id":
            return By.ID
        elif locatorType == "name":
            return By.NAME
        elif locatorType == "xpath":
            return By.XPATH
        elif locatorType == "css":
            return By.CSS_SELECTOR
        elif locatorType == "class":
            return By.CLASS_NAME
        elif locatorType == "link":
            return By.LINK_TEXT
        else:
            print("Locator type " + locatorType + " not correct/supported")
        return False



    def func_TestLogInitialization(self):
        try:
            self.config_data.iLogMsgSize = len(self.config_data.sLogMsg);
            self.config_data.iTCStatusSize = len(self.config_data.sTCStatus);
            self.config_data.iSCreenNamesSize = len(self.config_data.sSCreenNames);
            self.config_data.iStepExecutionDateSize = len(self.config_data.sStepExecutionDate);
            if (self.config_data.iCount == 0):
                self.config_data.sStartTime = self.func_GetTime();

            try :
             self.config_data.sReleaseName="";
            except  Exception as e:
             pass


            self.config_data.sSCreenNames = [];
            self.config_data.sLogMsg = [];
            self.config_data.sStepExecutionDate = [];
            sTCStatus = [];
            self.func_TCResultFolderCreation();
            self.initializeScrenCounter();

        except Exception as e:
            print("Exception:",e)
            pass
    def initializeScrenCounter(self):
        currentDirectory = os.path.dirname(__file__)
        parentDirectory = os.path.abspath(os.path.join(currentDirectory, os.pardir))
        parentDirectory = os.path.join(parentDirectory, self.config_data.screenshotDirectory,self.scriptName)
        
        
        if self.override_screenshot.lower()=="no":

             
            count = 1
            # Iterate directory
            for path in os.listdir(parentDirectory):
                # check if current path is a file
                if os.path.isfile(os.path.join(parentDirectory, path)):
                    count += 1
            self.config_data.iScreenCounter=count
    def func_TCResultFolderCreation(self):

        #os.path.join(self.config_data.projectpath,"Config Files","Basic_Setup.ini")
        destinationFile = os.path.join(self.config_data.projectpath, self.config_data.logDirectory,self.scriptName)
        screenshotDirectory = os.path.join(self.config_data.projectpath,self.config_data.screenshotDirectory,self.scriptName)
        self.config_data.screenshotDirectoryFull=screenshotDirectory
        if not os.path.isdir(destinationFile):
            os.makedirs(destinationFile)
        if not os.path.isdir(screenshotDirectory):
            os.makedirs(screenshotDirectory)
        self.config_data.sTestCaseLogFolder = str(destinationFile);

    def func_InformationMessageToPrintManualSteps(self, sScreenName='') :
        try :

            self.config_data.sLogMsg.append(sScreenName);
            self.config_data.sTCStatus.append("COMMENT");
            self.config_data.arSCreenNames.append("");
            self.config_data.iTotalScreenCounter.append(0);
            self.config_data.sStepExecutionDate.append(self.func_GetDate());

        except Exception as e:
            print(e)
            pass

    # Screen Validation Function
    def func_ScreenValidation(self,sScreenName):
            bResl = False
            sName=self.driver.title
            self.func_InformationMessageToPrintManualSteps("Validation for the screen: "+sScreenName);
            if sName.lower() == sScreenName.lower():

                self.func_ScreenCaptureWithPassMsg("Expected: " + sScreenName + ", Actual: "+sName);
                bResl = True;

            else:
              self.func_ScreenCaptureWithPassMsg("Expected: " + sScreenName + ", Actual: "+sName);

            return bResl;


    def func_ResetLogVariables(self):

        temp = [];
        for index in range(len(self.config_data.iLogMsgSize)-1):
            temp.append(self.config_data.sLogMsg.get(index));
        self.config_data.sLogMsg.clear();
        for index in range(len(self.config_data.iLogMsgSize)):
            self.config_data.sLogMsg.append(temp.get(index));
        temp.clear();
        for index in range(len(self.config_data.iTCStatusSize)-1):
            temp.append(self.config_data.sTCStatus.get(index));

        self.config_data.sTCStatus.clear();
        for index in range(len(temp)):
            self.config_data.sTCStatus.append(temp.get(index));
        temp.clear();

        for index in range(len(self.config_data.iSCreenNamesSize)-1):
            temp.append(self.config_data.sSCreenNames.get(index));

        self.config_data.sTCStatus.clear();
        for index in range(len(temp)):
            self.config_data.sSCreenNames.append(temp.get(index));
        temp.clear();

        for index in range(len(self.config_data.iStepExecutionDateSize)-1):
            temp.append(self.config_data.sStepExecutionDate.get(index));
        self.config_data.sTCStatus.clear();

        for index in range(len(temp)):
            self.config_data.sStepExecutionDate.append(temp.get(index));
        temp.clear();

    def func_SetEndTime(self):

        if "excel" in self.log_type.strip().lower():
            sDate = self.config_data.sDateInDDMMYYY;

            wb = load_workbook(filename=self.logPath)
            ws=wb.get_sheet_by_name(sDate)




            if self.config_data.sEndTime is not None:
                ws.cell(row=3, column=1).value = "End Time " + self.config_data.sEndTime
                try:
                    ws.cell(row=3, column=1).style = "log_border"
                except:
                    ws.cell(row=3, column=1).style = self.log_border;


            else:
                ws.cell(row=3, column=1).value = "End Time "+self.func_GetTime()
                try:
                    ws.cell(row=3, column=1).style = "log_border"
                except:
                    ws.cell(row=3, column=1).style = self.log_border;
            wb.save(self.logPath)

            wb = load_workbook(filename=self.consolidatedlogPath)
            ws = wb.get_sheet_by_name(sDate)
            if self.config_data.sEndTime is not None:
                ws.cell(row=3, column=1).value = "End Time " + self.config_data.sEndTime
                try:
                    ws.cell(row=3, column=1).style = "log_border"
                except:
                    ws.cell(row=3, column=1).style = self.log_border;

            else:
                ws.cell(row=3, column=1).value = "End Time "+self.func_GetTime()
                try:
                    ws.cell(row=3, column=1).style = "log_border"
                except:
                    ws.cell(row=3, column=1).style = self.log_border;
            wb.save(self.consolidatedlogPath)

    def func_ScreenCaptureWithPassMsg(self,sMsg):
         texttitle=''
         try:
             texttitle=self.driver.title

         except:
             pass
         try:

            if "log" in self.log_type.strip().lower() or "txt" in self.log_type.strip().lower():
                self.log.info("Screen Name: " + texttitle+ " : " + sMsg+". Screenshot Destination: "+self.func_ScreenCaptureFortxtLog())

            elif "excel" in self.log_type.strip().lower() or "html" in self.log_type.strip().lower():

                self.config_data.sLogMsg.append("Screen Name: " +texttitle + " : " + sMsg);
                self.config_data.sTCStatus.append("PASS");

                self.config_data.sSCreenNames.append(self.func_ScreenCapture());
                self.config_data.sStepExecutionDate.append(self.func_GetDate());



         except Exception as e:
             print("generic 244:",e)

    def func_SetLogMsgWithoutScreenCapture(self, sMsg):
        try:
            if "excel" in self.log_type.strip().lower() or "html" in self.log_type.strip().lower():
                self.config_data.sLogMsg.append(sMsg);
                self.config_data.sTCStatus.append("INFORMATION");
                self.config_data.arSCreenNames.append("");
                self.config_data.iTotalScreenCounter.append(0);
                self.config_data.sStepExecutionDate.append(self.func_GetDate());
            self.config_data.sTCWithStats[self.config_data.sTestcaseId] = "Pass";

        except Exception as e:
            pass

    def func_SetObjective(self, sMsg):
        try:
                self.config_data.sLogMsg.append(sMsg);
                self.config_data.sTCStatus.append("Objective");
                self.config_data.arSCreenNames.append("");
                self.config_data.iTotalScreenCounter.append(0);
                self.config_data.sStepExecutionDate.append(self.func_GetDate());
                self.config_data.sTCWithStats[self.config_data.sTestcaseId] = "Pass";

        except Exception as e:
            pass

    def func_SetTestcaseID(self,sMsg):
                self.config_data.sLogMsg.append(sMsg);
                self.config_data.sTCStatus.append("ID");
                self.config_data.arSCreenNames.append("");
                self.config_data.iTotalScreenCounter.append(0);
                self.config_data.sStepExecutionDate.append(self.func_GetDate());
                self.config_data.sTCWithStats[self.config_data.sTestcaseId] = "Pass";
    def func_PrintSetNo(self):
        if "txt" in self.log_type.strip().lower() or "log" in self.log_type.strip().lower():
            self.markTheEndOfTextlog("Test Case ID: " + self.config_data.sTestcaseId + "\n" + "-" * 25)

        if "excel" in self.log_type.strip().lower():
            self.func_SetLogMsgWithoutScreenCapture("Test Case ID: " + self.config_data.sTestcaseId);
        if "html" in self.log_type.strip().lower():
            self.func_SetTestcaseID("-"*20+"Test Case ID: " + self.config_data.sTestcaseId+"-"*20);

    def func_ScreenCaptureWithFailMsg(self,sMsg):
        texttitle = ''
        try:
            texttitle = self.driver.title
        except:
            pass

        try:
            if "log" in self.log_type.strip().lower() or "txt" in self.log_type.strip().lower():
                self.log.error("Screen Name: " + texttitle+ " : " + sMsg+". Screenshot Destination: "+self.func_ScreenCaptureFortxtLog())

            elif "excel" in self.log_type.strip().lower() or "html" in self.log_type.strip().lower():
                self.config_data.sLogMsg.append("Screen Name: " + texttitle+ " : " + sMsg);
                self.config_data.sTCStatus.append("FAIL");
                self.config_data.sSCreenNames.append(self.func_ScreenCapture());
                self.config_data.sStepExecutionDate.append(self.func_GetDate());

        except Exception as e:
               # TODO: handle exception
            pass



    def func_LogBodyCreation(self):
        if "txt" in self.log_type.strip().lower() or "log" in self.log_type.strip().lower():
            self.markTheEndOfTextlog("=" * 140)

        if "html" in self.log_type.strip().lower():
            currentDirectory = os.path.dirname(__file__)
            logpath=os.path.join(currentDirectory,self.config_data.logDirectory,self.config_data.sNewTestCaseName, self.config_data.sDateInDDMMYYY)
            
            if not os.path.isdir(logpath):
                os.makedirs(logpath)

            logfile=os.path.join(logpath,self.config_data.sNewTestCaseName+".html")
            self.generateHtmlLog(logfile,self.config_data.sLogMsg,"")
        if "excel" in self.log_type.strip().lower():
            sDate =self.config_data.sDateInDDMMYYY;
            bRes = True;
            bCount = True;
            bTCCount = True;
            iStatus = 0;

            self.config_data.sLogFilePath=self.logPath
            if (len(self.config_data.sLogMsg) != 0):

                wb = load_workbook(filename=self.logPath)
                ws = wb.get_sheet_by_name(sDate)
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 30
                ws.column_dimensions['E'].width = 60
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 20
                headers = ["Date and Time of Execution", "Release Name", "Testcase Name",
                           "Status",
                           "Comments", "Screen Capture", "Input Data"]
                for index, data in enumerate(headers):

                    ws.cell(row=6, column=index + 1).value = data
                    ws.cell(row=6, column=index + 1).alignment = Alignment(wrap_text=True)
                    Logheader = NamedStyle(name="Logheader")
                    bd = Side(border_style='thin')
                    border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    Logheader.border = border
                    Logheader.alignment = center_alignment
                    Logheader.fill = PatternFill(start_color="FF7F3F", fill_type="solid")
                    Logheader.alignment.wrap_text = True
                    try:
                        ws.cell(row=6, column=index + 1).style = "Logheader";

                    except Exception as e:
                        ws.cell(row=6, column=index + 1).style = Logheader;

                row_count = ws.max_row
                row_count+=1

                ws.cell(row=2, column=1).value = "Start Time:"+self.config_data.sStartTime
                try:
                    ws.cell(row=2, column=1).style = "log_border"
                except:
                    ws.cell(row=2, column=1).style = self.log_border;

                for i,data in enumerate(self.config_data.sLogMsg):

                    ws.cell(row=row_count, column=1).value =self.config_data.sStepExecutionDate[i]
                    try:
                        ws.cell(row=row_count, column=1).style = "log_border"
                    except:
                        ws.cell(row=row_count, column=1).style = self.log_border;
                    ws.cell(row=row_count, column=2).value = "Release:"+self.config_data.sReleaseName
                    try:
                        ws.cell(row=row_count, column=2).style = "log_border"
                    except:
                        ws.cell(row=row_count, column=2).style = self.log_border;
                    ws.cell(row=row_count, column=3).value =  self.config_data.sNewTestCaseName
                    try:
                        ws.cell(row=row_count, column=3).style = "log_border"
                    except:
                        ws.cell(row=row_count, column=3).style = self.log_border;
                    if  "Test Case ID:" in data:
                        self.config_data.iTotalTestCasesToRun+=1
                        bTCCount = True;
                    if "PASS" in self.config_data.sTCStatus[i]:
                        ws.cell(row=row_count, column=4).value = "PASS"
                        try:
                            ws.cell(row=row_count, column=4).style = "log_pass"
                        except:
                            ws.cell(row=row_count, column=4).style = self.log_pass;

                    elif "INFORMATION" in self.config_data.sTCStatus[i]:
                        ws.cell(row=row_count, column=4).value = "INFORMATION"

                        try:
                            ws.cell(row=row_count, column=4).style = "log_info"
                        except:
                            ws.cell(row=row_count, column=4).style = self.log_info;
                    elif "COMMENT" in self.config_data.sTCStatus[i]:
                        ws.cell(row=row_count, column=4).value = "COMMENT"
                        try:
                            ws.cell(row=row_count, column=4).style = "log_comment"
                        except:
                            ws.cell(row=row_count, column=4).style = self.log_comment;
                    elif "FAIL" in self.config_data.sTCStatus[i]:
                        ws.cell(row=row_count, column=4).value = "FAIL"
                        self.config_data.sTCWithStats[self.config_data.sNewTestCaseName]="FAIL"
                        try:
                            ws.cell(row=row_count, column=4).style = "log_fail"
                        except:
                            ws.cell(row=row_count, column=4).style = self.log_fail;

                        if bTCCount:
                            self.config_data.iTotalTestCasesFail+=1
                            bTCCount=False






                    if self.config_data.iTotalScreenCounter[i] !=0:
                        hypertxt=self.func_GetHyperText(self.config_data.iTotalScreenCounter[i])
                        ws.cell(row=row_count, column=6).value = hypertxt
                        try:
                            ws.cell(row=row_count, column=6).style = "log_border"
                        except:
                            ws.cell(row=row_count, column=6).style = self.log_border;
                    if bCount:
                        hypertxt=self.func_GetHyperTextForInputExcel()

                        ws.cell(row=row_count, column=7).value = hypertxt
                        try:
                            ws.cell(row=row_count, column=7).style = "log_border"
                        except:
                            ws.cell(row=row_count, column=7).style = self.log_border;

                    if "objective" in data.lower() or "Test Case ID".lower() in data.lower():
                        ws.cell(row=row_count, column=5).value = data

                        try:
                            ws.cell(row=row_count, column=5).style = "log_set"
                        except:
                            ws.cell(row=row_count, column=5).style = self.log_set

                    else:
                        ws.cell(row=row_count, column=5).value = data
                        try:
                            ws.cell(row=row_count, column=5).style = "log_border"
                        except:
                            ws.cell(row=row_count, column=5).style = self.log_border;
                    row_count+=1
                self.config_data.iTotalTestCasesPass = self.config_data.iTotalTestCasesToRun- self.config_data.iTotalTestCasesFail;



                wb.save(self.logPath)

                self.func_ConsolidatedLogBodyCreation()


    def func_ConsolidatedLogBodyCreation(self):

        sDate = self.config_data.sDateInDDMMYYY;
        print("=============sDate",sDate)
        print("===========self.consolidatedlogPath:",self.consolidatedlogPath)
        wb_con = load_workbook(filename=self.consolidatedlogPath)
        ws_con = wb_con.get_sheet_by_name(sDate)
        ws_con.column_dimensions['A'].width = 30
        ws_con.column_dimensions['B'].width = 20
        ws_con.column_dimensions['C'].width = 30
        ws_con.column_dimensions['D'].width = 20
        ws_con.column_dimensions['E'].width = 20
        row_count = ws_con.max_row
        row_count+=1
        sObjective = "";
        if (len(self.config_data.sTCIDWithObjective)== 1) :
            for key in self.config_data.sTCIDWithObjective:
                sObjective = self.config_data.sTCIDWithObjective[key]


        icounter = 0;

        for key in self.config_data.sTCWithStats:
            ws_con .cell(row=row_count, column=1).value = self.config_data.sStepExecutionDate[icounter];
            try:
                ws_con.cell(row=row_count, column=1).style = "log_border"
            except:
                ws_con.cell(row=row_count, column=1).style = self.log_border;
            ws_con.cell(row=row_count, column=2).value = "Release:" +self.config_data.sReleaseName;
            try:
                ws_con.cell(row=row_count, column=2).style = "log_border"
            except:
                ws_con.cell(row=row_count, column=2).style = self.log_border;
            ws_con.cell(row=row_count, column=3).value = self.config_data.sNewTestCaseName;
            try:
                ws_con.cell(row=row_count, column=3).style = "log_border"
            except:
                ws_con.cell(row=row_count, column=3).style = self.log_border;
            if len(self.config_data.sTCIDWithObjective)==1:
                ws_con.cell(row=row_count, column=4).value = sObjective;
                try:
                    ws_con.cell(row=row_count, column=4).style = "log_border"
                except:
                    ws_con.cell(row=row_count, column=4).style = self.log_border;
            else:
                ws_con.cell(row=row_count, column=4).value = self.config_data.sTCIDWithObjective[key];
                try:
                    ws_con.cell(row=row_count, column=4).style = "log_border"
                except:
                    ws_con.cell(row=row_count, column=4).style = self.log_border;
            if "pass" in self.config_data.sTCWithStats[key].lower():
                ws_con.cell(row=row_count, column=5).value = "PASS"

                try:
                    ws_con.cell(row=row_count, column=5).style = "log_pass"
                except:
                    ws_con.cell(row=row_count, column=5).style = self.log_pass;

            elif "fail" in self.config_data.sTCWithStats[key].lower():
                ws_con.cell(row=row_count, column=5).value = "FAIL"
                try:
                    ws_con.cell(row=row_count, column=5).style = "log_fail"
                except:
                    ws_con.cell(row=row_count, column=5).style = self.log_fail;

            row_count+=1

        ws_con.cell(row=1,
                column=1).value = "Test Log Report of " + self.config_data.sLogHeader + " as on  " + self.config_data.sDateInDDMMYYY
        ws_con.cell(row=1,column=1).alignment = Alignment(wrap_text=True)

        ws_con.cell(row=2, column=1).value = "Start Time:" + self.config_data.sStartTime
        try:
            ws_con.cell(row=2, column=1).style = "log_border"
        except:
            ws_con.cell(row=2, column=1).style = self.log_border;
        ws_con.cell(row=2, column=2).value = self.config_data.iTotalTestCasesPass;
        try:
            ws_con.cell(row=2, column=2).style = "log_pass"
        except:
            ws_con.cell(row=2, column=2).style = self.log_pass;
        ws_con.cell(row=2, column=3).value = self.config_data.iTotalTestCasesFail
        try:
            ws_con.cell(row=2, column=3).style = "log_fail"
        except:
            ws_con.cell(row=2, column=3).style = self.log_fail;
        ws_con.cell(row=2, column=4).value = self.config_data.iTotalTestCasesToRun
        try:
            ws_con.cell(row=2, column=4).style = "log_border";
        except:
            ws_con.cell(row=2, column=4).style = self.log_border;
        header2 = ["Total No of Test scripts Executed", "Total No of Test scripts Pass",
                   "Total No of Test scripts Fail"]
        for index, data in enumerate(header2):

            ws_con.cell(row=1, column=index + 2).value = data
            try:
                ws_con.cell(row=1, column=index+2).style = "Logheader";
            except:
                ws_con.cell(row=1, column=index+2).style = self.Logheader;
        wb_con.save(self.consolidatedlogPath)


    def func_GetHyperText(self,iScreenCounter):
        sHyperlink = "=HYPERLINK(\"" \
                 + self.config_data.sNewTestCaseName + "\\" \
                 + self.config_data.sCountry + "\\" \
                 + self.config_data.sDateInDDMMYYY + "\\" \
                 + self.config_data.sLogScreenshot + str(iScreenCounter) \
                 + ".jpg\",\"" + self.config_data.sLogScreenshotAlias \
                 + str(iScreenCounter) + "\")";

        return sHyperlink;

    def func_GetHyperTextForInputExcel(self):

        sHyperlink = "=HYPERLINK(\"" \
                 + self.config_data.sInputFileName + "#'" \
                 + self.config_data.sProcessSheetName + "'!A" \
                 + (str(self.config_data.iTestcaseIDRowNo + 1)) \
                 + "\",\"Input Datasheet\")";

        return sHyperlink;


    def func_GetDate(self):

        dateString=datetime.today().strftime('%B %d %Y, %A, %I:%M %p')
        return dateString;
    def func_GetTime(self):

        dateString=datetime.today().strftime('%I:%M %p')
        return dateString;
    def    func_GetDateInDDMMYYYYFormat(self):

        dateString = datetime.today().strftime('%d%m%Y')

        return dateString;

    # It creates a  blank template in the Log file.
    def    func_LogHeaderCreation(self):

        try :

            self.config_data.sDateInDDMMYYY = self.func_GetDateInDDMMYYYYFormat();
            sDate = self.config_data.sDateInDDMMYYY;

            self.logPath = os.path.join(self.config_data.sLogFilePath, self.config_data.sScriptFolderName, sDate,
                                        self.config_data.sLogFileName_excel)
            self.logDirectory = os.path.join(self.config_data.sLogFilePath, self.config_data.sScriptFolderName, sDate)

            if "excel" in self.log_type.strip().lower():


                    bRes = True;
                    try :

                        headers = ["Date and Time of Execution", "Release Name", "Testcase Name",
                                   "Status",
                                   "Comments", "Screen Capture", "Input Data"]

                        existingFlag=True
                        if not os.path.exists(self.logDirectory):

                            os.makedirs(self.logDirectory)
                        if os.path.exists(self.logPath):


                            sheets=pd.ExcelFile(self.logPath).sheet_names
                            wb=None
                            if sDate not in sheets:
                                wb = load_workbook(filename=self.logPath)

                                ws=wb.create_sheet(sDate, 0);


                                ws.cell(row=1,column=1).value = "Test Log Report of " + self.config_data.sLogHeader + " as on  " + self.config_data.sDateInDDMMYYY
                                ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
                                try:
                                    ws.cell(row=1, column=1).style = "log_border"
                                except:
                                    ws.cell(row=1, column=1).style = self.log_border;


                                if (self.config_data.iCount == 0):
                                    self.config_data.iCount = 1;
                                    if self.config_data.sStartTime is not None:
                                        ws.cell(row=2, column=1).value ="Start Time " + self.config_data.sStartTime
                                    else:
                                        ws.cell(row=2, column=1).value = "Start Time "
                                    try:
                                        ws.cell(row=2, column=1).style = "log_border"
                                    except:
                                        ws.cell(row=2, column=1).style = self.log_border;

                                if self.config_data.sEndTime is not None:
                                    ws.cell(row=3, column=1).value="End Time " + self.config_data.sEndTime

                                    try:
                                        ws.cell(row=3, column=1).style = "log_border"
                                    except:
                                        ws.cell(row=3, column=1).style = self.log_border;

                                else:
                                    ws.cell(row=3, column=1).value="End Time "
                                    try:
                                        ws.cell(row=3, column=1).style = "log_border"
                                    except:
                                        ws.cell(row=3, column=1).style = self.log_border;


                                ws.cell(row=4, column=1).style = "Logheader"
                                ws.cell(row=4, column=1).value = "Execution Details"




                                wb.save(self.logPath)

                            else:
                                try:

                                    excel_df = pd.read_excel(self.logPath,sheet_name=sDate);
                                    self.excel.lastRow=excel_df.shape[0]


                                except:
                                    pass

                            existingFlag = False
                        else:

                            excel_df=pd.DataFrame(columns=headers)

                            self.excel.logwriter = pd.ExcelWriter(self.logPath, engine='xlsxwriter')

                            # Convert the dataframe to an XlsxWriter Excel object.
                            excel_df.to_excel(self.excel.logwriter, sheet_name=sDate, startrow=5, startcol=0,header=True, index=False)
                            # Get the xlsxwriter objects from the dataframe writer object.
                            self.logWorkbook = self.excel.logwriter.book
                            self.logWorksheet = self.excel.logwriter.sheets[sDate]

                        if existingFlag:

                            log_border_format = self.logWorkbook.add_format({'border': 1, 'valign': 'top'})
                            self.log_border_format=log_border_format
                            self.logWorksheet.write(0,0,"Test Log Report of " + self.config_data.sLogHeader + " as on  "
                            + self.config_data.sDateInDDMMYYY,self.log_border_format)

                            if (self.config_data.iCount == 0):
                                self.config_data.iCount = 1;
                                try:
                                    self.logWorksheet.write(1,0,"Start Time " + self.config_data.sStartTime, self.log_border_format)
                                except:
                                    self.logWorksheet.write(1, 0, "Start Time " + self.config_data.sStartTime,
                                                            'log_border_format')


                            if self.config_data.sEndTime is not None:
                                try:
                                    self.logWorksheet.write( 2,0, "End Time "+self.config_data.sEndTime, self.log_border_format)
                                except:
                                    self.logWorksheet.write( 2,0, "End Time "+self.config_data.sEndTime, 'log_border_format')


                            else:
                                try:
                                    self.logWorksheet.write(2, 0, "End Time ", self.log_border_format)
                                except:
                                    self.logWorksheet.write(2, 0, "End Time ", 'log_border_format')


                            self.Logheader = self.logWorkbook.add_format({'border': 1, 'bold': True, 'valign': 'top'})
                            self.Logheader.set_align('vcenter')
                            self.Logheader.set_bg_color('#FF7F3F')

                            for index,data in enumerate(headers):
                                self.logWorksheet.write(5, index, data,self.Logheader)
                    except Exception as  e :
                        print("Exception:",e)
                        pass

                    if self.excel.logwriter is not None:
                            self.excel.logwriter.save()


                    #self.config_data.sLogScreenshot = "selenium_";
                    self.config_data.sLogScreenshot = "";
                    self.config_data.sLogScreenshotAlias = "Screen";


        except Exception as e1:
                print("Exception1:",e1)
                pass




    def func_ConsolidatedLogHeaderCreation(self):
        if "excel" in self.log_type.strip().lower():
                headers = ["Date and Time of Execution", "Release Name", "Testcase Name", "Objective", "Status"]
                header2 = ["Total No of Test scripts Executed", "Total No of Test scripts Pass",
                           "Total No of Test scripts Fail"]
                if self.config_data.sDateInDDMMYYY is not None:
                    sDate = self.config_data.sDateInDDMMYYY;
                else:
                    sDate=self.func_GetDateInDDMMYYYYFormat();
                    self.config_data.sDateInDDMMYYY=sDate;
                try:

                    bRes = True;
                    try:
                        self.consolidatedlogPath = os.path.join(self.config_data.sLogFilePath, self.config_data.sConsolidatedLogFileName)
                        existingFlag = True
                        if os.path.isfile(self.consolidatedlogPath):

                            sheets = pd.ExcelFile(self.consolidatedlogPath).sheet_names
                            wb = None
                            if sDate not in sheets:

                                wb = load_workbook(filename=self.consolidatedlogPath)
                                # Get the current Active Sheet
                                # ws = wb.get_active_sheet()

                                # ws = wb.get_sheet_by_name(sDate)
                                ws = wb.create_sheet(sDate, 0);


                                ws.cell(row=1,column=1).value = "Test Log Report of " + self.config_data.sLogHeader + " as on  " + self.config_data.sDateInDDMMYYY
                                ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
                                try:
                                    ws.cell(row=1, column=1).style = "con_log_border"
                                except:
                                    ws.cell(row=1, column=1).style = self.con_log_border
                                if (self.config_data.iCount == 0):
                                    self.config_data.iCount = 1;
                                    if self.config_data.sStartTime is not None:
                                        ws.cell(row=2, column=1).value = "Start Time " + self.config_data.sStartTime
                                    else:
                                        ws.cell(row=2, column=1).value = "Start Time "
                                    ws.cell(row=2, column=1).style = "con_log_border"
                                if self.config_data.sEndTime is not None:
                                    ws.cell(row=3, column=1).value = "End Time " + self.config_data.sEndTime
                                    ws.cell(row=3, column=1).style = "con_log_border"

                                else:
                                    ws.cell(row=3, column=1).value = "End Time "
                                    ws.cell(row=3, column=1).style = "con_log_border"




                                for index, data in enumerate(headers):
                                    try:
                                        ws.cell(row=5, column=index + 1).style = "Logheader"
                                    except:
                                        ws.cell(row=5, column=index + 1).style = self.Logheader
                                    ws.cell(row=5, column=index + 1).value = data
                                wb.save(self.consolidatedlogPath)

                            else:
                                try:
                                    excel_df = pd.read_excel(self.consolidatedlogPath, sheet_name=sDate);
                                    excel_df=excel_df.fillna("")
                                    self.excel.lastRow=excel_df


                                except:
                                    pass

                            existingFlag = False
                        else:

                            excel_df = pd.DataFrame(columns=headers)

                            self.excel.consolidate_logwriter = pd.ExcelWriter(self.consolidatedlogPath, engine='xlsxwriter')

                            # Convert the dataframe to an XlsxWriter Excel object.
                            excel_df.to_excel(self.excel.consolidate_logwriter, sheet_name=sDate, startrow=5, startcol=0,
                                              header=True, index=False)
                            print("self.excel.logwriter",self.excel.logwriter)
                            # Get the xlsxwriter objects from the dataframe writer object.
                            self.logWorkbook = self.excel.logwriter.book
                            self.logWorksheet = self.excel.logwriter.sheets[sDate]

                        if existingFlag:

                            self.log_border_format = self.logWorkbook.add_format({'border': 1, 'valign': 'top'})
                            self.logWorksheet.write(0, 0, "Test Log Report of " + self.config_data.sLogHeader + " as on  "
                                                    + sDate, self.log_border_format)

                            if (self.config_data.iCount == 0):
                                self.config_data.iCount = 1;
                                if self.config_data.sStartTime is not None:
                                    self.logWorksheet.write(1, 0, "Start Time " + self.config_data.sStartTime,
                                                        self.log_border_format)
                                else:
                                    self.logWorksheet.write(1, 0, "Start Time ",
                                                            self.log_border_format)
                            if self.config_data.sEndTime is not None:

                                self.logWorksheet.write(2, 0, "End Time " + self.config_data.sEndTime,
                                                        self.log_border_format)
                            else:
                                self.logWorksheet.write(2, 0, "End Time ", self.log_border_format)

                            self.Logheader = self.logWorkbook.add_format({'border': 1, 'bold': True, 'valign': 'top'})
                            self.Logheader.set_align('vcenter')
                            self.Logheader.set_bg_color('#FF7F3F')
                            for index, data in enumerate(header2):
                                self.logWorksheet.write(0, index+1, data, self.Logheader)
                            for index, data in enumerate(headers):
                                self.logWorksheet.write(5, index, data, self.Logheader)
                    except Exception as e:
                        print("Exception>", e)
                        pass
                    try:
                        if self.excel.consolidate_logwriter is not None:
                            self.excel.consolidate_logwriter.save()
                    except:
                        pass

                    #self.config_data.sLogScreenshot = "selenium_";
                    self.config_data.sLogScreenshot = "";
                    self.config_data.sLogScreenshotAlias = "Screen";


                except Exception as e1:
                    print("Exception1:",e1)
                    pass
                finally:
                    pass




    def   func_ResultValidation(self,actualRes,expectedRes):

        if (expectedRes in actualRes):
            self.func_ScreenCaptureWithPassMsg("Actual Result: " + actualRes + " , Expected Result: " + expectedRes);

        else:

            self.func_ScreenCaptureWithFailMsg("Actual Result: " + actualRes + " , Expected Result: " + expectedRes);

    def func_ScreenCaptureFortxtLog(self):
            """
            Takes screenshot of the current open web page
            """
            screenshotDirectory = "Logs/Screenshots/"
            currentDirectory = os.path.dirname(__file__)
            parentDirectory=os.path.abspath(os.path.join(currentDirectory, os.pardir))
            parentDirectory=os.path.join(parentDirectory,screenshotDirectory)
            fileName=os.path.join(parentDirectory, str(self.config_data.sLogScreenshot) + str(self.config_data.iScreenCounter) + ".png")
            print("fileName"+fileName)
            try:
                if not os.path.exists(parentDirectory):
                    os.makedirs(parentDirectory)
                self.driver.save_screenshot(fileName)

            except Exception as e:
                if (self.log != None):

                    self.log.error("### Exception Occurred when taking screenshot:"+e)
                pass

            self.config_data.iScreenCounter = self.config_data.iScreenCounter + 1;
            #     Utility.self.config_data.sTestCaseLogFolder + "\\" + Utility.self.config_data.sLogScreenshot + Utility.self.config_data.iScreenCounter + ".jpg"));
            self.config_data.iTotalScreenCounter.append(self.config_data.iScreenCounter);

            return "Logs/Screenshots/"+str(self.config_data.sLogScreenshot) +str(self.config_data.iScreenCounter) + ".jpg"
    def func_ScreenCapture(self):
            """
            Takes screenshot of the current open web page
            """
            screenshotDirectory = "Logs/Screenshots/"
            currentDirectory = os.path.dirname(__file__)
            parentDirectory = os.path.abspath(os.path.join(currentDirectory, os.pardir))
            
            parentDirectory = os.path.join(parentDirectory, screenshotDirectory,self.scriptName)
            print("parentDirectory",parentDirectory)
            print("=================str(self.config_data.sLogScreenshot)",self.scriptName)
            fileName = os.path.join(parentDirectory, self.scriptName + str(
                self.config_data.iScreenCounter) + ".png")
            print("=========file:",fileName)
            try:
                if not os.path.exists(parentDirectory):
                    os.makedirs(parentDirectory)
                self.driver.save_screenshot(fileName)

            except Exception as e:
                if (self.log != None):

                    self.log.error("### Exception Occurred when taking screenshot:" + e)
                pass
            self.config_data.iTotalScreenCounter.append(self.config_data.iScreenCounter);

            self.config_data.iScreenCounter = self.config_data.iScreenCounter + 1;
            #     Utility.self.config_data.sTestCaseLogFolder + "\\" + Utility.self.config_data.sLogScreenshot + Utility.self.config_data.iScreenCounter + ".jpg"));

            return self.scriptName +str( self.config_data.iScreenCounter-1);


    def func_ScreenCaptureWithoutMsgFortxtLog(self,sMsg):

        try:

            self.config_data.sLogMsg.append("");
            self.config_data.sSCreenNames.append(self.func_ScreenCaptureFortxtLog());
            self.config_data.iTotalScreenCounter.append(self.config_data.iScreenCounter);
            self.config_data.sStepExecutionDate.append(self.func_GetDate());


        except Exception as e :

               # TODO: handle exception
            pass

    def func_ScreenCaptureWithoutMsg(self,sMsg):

        try:

            self.config_data.sLogMsg.append("");
            self.config_data.sSCreenNames.append(self.func_ScreenCapture());
            self.config_data.iTotalScreenCounter.append(self.config_data.iScreenCounter);
            self.config_data.sStepExecutionDate.append(self.func_GetDate());


        except Exception as e :

               # TODO: handle exception
            pass

    def getElement(self, locator, locatorType="id"):
        element = None
        try:
            locatorType = locatorType.lower()
            byType = self.getByType(locatorType)
            element = self.driver.find_element(byType, locator)
            print("Element Found with locator: " + locator + " and  locatorType: " + locatorType)
        except:
            print("Element not found with locator: " + locator + " and  locatorType: " + locatorType)
        return element

    def elementClick(self, locator, locatorType="id"):
        try:
            element = self.getElement(locator, locatorType)


            element.click()
            print("Clicked on element with locator: " + locator + " locatorType: " + locatorType)
        except:
            print("Cannot click on the element with locator: " + locator + " locatorType: " + locatorType)
            print_stack()

    def sendKeys(self, data, locator, locatorType="id"):
        try:
            element = self.getElement(locator, locatorType)
            element.send_keys(data)
            print("Sent data on element with locator: " + locator + " locatorType: " + locatorType)
        except:
            print("Cannot send data on the element with locator: " + locator +
                  " locatorType: " + locatorType)
            print_stack()

    def isElementPresent(self, locator, locatorType="id"):
        try:
            element = self.getElement(locator, locatorType)
            if element is not None:
                print("Element Found")
                return True
            else:
                print("Element not found")
                return False
        except:
            print("Element not found")
            return False

    def elementPresenceCheck(self, locator, byType):
        try:
            elementList = self.driver.find_elements(byType, locator)
            if len(elementList) > 0:
                print("Element Found")
                return True
            else:
                print("Element not found")
                return False
        except:
            print("Element not found")
            return False

    def waitForElement(self, locator, locatorType="id",timeout=10, pollFrequency=0.5):
        element = None
        try:
            byType = self.getByType(locatorType)
            print("Waiting for maximum :: " + str(timeout) +
                  " :: seconds for element to be clickable")
            wait = WebDriverWait(self.driver, timeout, poll_frequency=pollFrequency,
                                 ignored_exceptions=[NoSuchElementException,
                                                     ElementNotVisibleException,
                                                     ElementNotSelectableException])
            element = wait.until(EC.element_to_be_clickable((byType, locator)))
            print("Element appeared on the web page")
        except:
            print("Element not appeared on the web page")
            print_stack()
        return element
    def setText(self, locator, value):
        if str(locator).endswith("_XPATH"):

            self.driver.find_element_by_xpath(ConfigReader.readConfig("locators", locator)).send_keys(value)
        elif str(locator).endswith("_CSS"):
            self.driver.find_element_by_css_selector(ConfigReader.readConfig("locators", locator)).send_keys(value)
        elif str(locator).endswith("_ID"):
            self.driver.find_element_by_id(ConfigReader.readConfig("locators", locator)).send_keys(value)

    def checkbox_Select(self, field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)

        self.func_InformationMessageToPrintManualSteps("Check Box "+field+" is checked");

        if self.driver.find_element(By.XPATH, locator_details[0]).is_selected()==False:
            self.driver.find_element(By.XPATH, locator_details[0]).click()
        if (self.log != None):

           self.log.logger.info("Select the Checkbox: " + str(field))
    def checkbox_UnSelect(self, field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)

        self.func_InformationMessageToPrintManualSteps("Check Box "+field+" is unchecked");

        if self.driver.find_element(By.XPATH, locator_details[0]).is_selected() == True:
            self.driver.find_element(By.XPATH, locator_details[0]).click()

        self. log.logger.info("UnSelect the Checkbox: " + str(field))

    def validate_checkbox_Checked(self,field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)
        if self.driver.find_element(By.XPATH, locator_details[0]).is_selected() == True:
            self.func_ScreenCaptureWithPassMsg("Check Box "+field +" is Selected")
        else:
            self.func_ScreenCaptureWithFailMsg("Check Box " + field + " is Not Selected")
    def validate_checkbox_Unchecked(self,field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)
        if self.driver.find_element(By.XPATH, locator_details[0]).is_selected() == False:
            self.func_ScreenCaptureWithPassMsg("Check Box " + field + " is unselected")
        else:
            self.func_ScreenCaptureWithFailMsg("Check Box " + field + " is selected")
    def validate_Field(self,field,text_validate):
        self.func_InformationMessageToPrintManualSteps("Fieled Validation")
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)

        expected=self.driver.find_element(By.XPATH, locator_details[0]).text

        if  expected== text_validate:

            self.func_ScreenCaptureWithPassMsg("Expected Field Name: " + text_validate +" , Actual Field Name: " + expected )
        else:
            self.func_ScreenCaptureWithFailMsg("Expected Field Name: " + text_validate +" , Actual Field Name: " + expected )
    def radiobutton_Select(self, field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)

        self.func_InformationMessageToPrintManualSteps("Radio Button "+field+" is checked");

        if self.driver.find_element(By.XPATH, locator_details[0]).is_selected()==False:
            self.driver.find_element(By.XPATH, locator_details[0]).click()
        if (self.log != None):

              self.log.logger.info("Check the Radio Button: " + str(field))
    def radiobutton_UnSelect(self, field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)

        self.func_InformationMessageToPrintManualSteps("Radio Button "+field+" is unchecked");

        if self.driver.find_element(By.XPATH, locator_details[0]).is_selected() == True:
            self.driver.find_element(By.XPATH, locator_details[0]).click()
        if (self.log != None):

              self.log.logger.info("Uncheck the Radio Button: " + str(field))

    def validate_radiobutton_Checked(self,field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)
        if self.driver.find_element(By.XPATH, locator_details[0]).is_selected() == True:
            self.func_ScreenCaptureWithPassMsg("Radio Button "+field +" is Selected")
        else:
            self.func_ScreenCaptureWithFailMsg("Radio Button " + field + " is Not Selected")
    def validate_radiobutton_Unchecked(self,field):
        title = self.driver.title.strip()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]
        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)
        if self.driver.find_element(By.XPATH, locator_details[0]).is_selected() == False:
            self.func_ScreenCaptureWithPassMsg("Radio Button " + field + " is unselected")
        else:
            self.func_ScreenCaptureWithFailMsg("Radio Button " + field + " is selected")
    def moveTo(self, locator):

        if str(locator).endswith("_XPATH"):
            element = self.driver.find_element_by_xpath(ConfigReader.readConfig("locators", locator))
        elif str(locator).endswith("_CSS"):
            element = self.driver.find_element_by_css_selector(ConfigReader.readConfig("locators", locator))
        elif str(locator).endswith("_ID"):
            element = self.driver.find_element_by_id(ConfigReader.readConfig("locators", locator))

        action = ActionChains(self.driver)
        action.move_to_element(element).perform()
        if (self.log != None):
            self.log.logger.info("Moving to an element: " + str(locator))
    def click(self, locator):
        self.driver.find_element(By.XPATH, locator).click()

        if (self.log != None):
            self.log.logger.info("Clicking on an element: " + str(locator))

    def set(self, locator, value):

        self.driver.find_element(By.XPATH, locator).send_keys(value)


        if(self.log!=None):
            self.log.logger.info("Typing in an element: " + str(locator) + " value entered as : " + str(value))

    def select(self, locator, value):
        drpdwn=self.driver.find_element(By.XPATH, locator)

        select = Select(drpdwn)
        select.select_by_visible_text(value)
        if (self.log != None):
            self.log.logger.info("Selecting from an element: " + str(locator) + " value selected as : " + str(value))
    def getRowsCountInATable(self, field):
        try:
            title = self.driver.title.strip()
            self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
            locator_details = []
            locators = self.dict_screen_details[field]
            tag = None
            for data in locators:
                if "tag_name:" in data:
                    tag = data.replace("tag_name:", "")
                else:
                    locator_details.append(data)
            table=self.driver.find_elements(By.XPATH, locator_details[0]+"/tbody/tr")
            # identifying the number of rows having <tr> tag

            # len method is used to get the size of that list
            self.func_InformationMessageToPrintManualSteps("Number of Rows In the Table: "+str(len(table)-1))
            return len(table);
        except Exception as e:
            print(e)
    def getHeaderCountInATable(self, field):
        try:
            title = self.driver.title.strip()
            self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
            locator_details = []
            locators = self.dict_screen_details[field]
            tag = None
            for data in locators:
                if "tag_name:" in data:
                    tag = data.replace("tag_name:", "")
                else:
                    locator_details.append(data)
            table=self.driver.find_elements(By.XPATH, locator_details[0]+"/tbody/tr[1]/th")

            self.func_InformationMessageToPrintManualSteps("Number of Rows In the Table: "+str(len(table)))
            return len(table);
        except Exception as e:
            print(e)
    def getHeaderNamesOfATable(self, field):
        try:
            title = self.driver.title.strip()
            self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
            locator_details = []
            locators = self.dict_screen_details[field]
            tag = None
            for data in locators:
                if "tag_name:" in data:
                    tag = data.replace("tag_name:", "")
                else:
                    locator_details.append(data)
            table=self.driver.find_elements(By.XPATH, locator_details[0]+"/tbody/tr[1]/th")
            names=''
            for h in table:
                names=names+","+h.text.strip()
            names=names[1:]
            self.func_InformationMessageToPrintManualSteps("Table Header Names: "+names)
            return names;
        except Exception as e:
            print(e)
    def validateAlertMessage(self,msg):
        alert = self.driver.switch_to.alert
        txtmsg = alert.text
        txtmsg = txtmsg.strip()
        if msg.strip() == txtmsg:
            self.func_ScreenCaptureWithPassMsg("Expected Message: " + msg.strip() + " Actual Message: " + txtmsg)
        else:
            self.func_ScreenCaptureWithFailMsg("Expected Message: " + msg.strip() + " Actual Message: " + txtmsg)
        try:
            alert.accept()
        except:
            pass
    def validateConfirmationMessage(self,msg):
        alert = self.driver.switch_to.alert
        txtmsg=alert.text
        txtmsg=txtmsg.strip()
        if msg.strip()== txtmsg:
            self.func_ScreenCaptureWithPassMsg("Expected Message: "+msg.strip()+" Actual Message: "+txtmsg)
        else:
            self.func_ScreenCaptureWithFailMsg("Expected Message: " + msg.strip()+" Actual Message: " + txtmsg)
        try:
            alert.dismiss()
        except:
            pass

    def getDataFromSpecific_Row_Col_FromTable(self,row,col,field):
        data=self.getAllDataFromTable(field)
        self.func_InformationMessageToPrintManualSteps("Data From the Row:"+row+" , Col:"+col+" is:"+data[row-1][col-1])
    def getAllDataFromTable(self, field):
        try:

            title = self.driver.title.strip()
            self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
            locator_details = []
            locators = self.dict_screen_details[field]
            tag = None
            for data in locators:
                if "tag_name:" in data:
                    tag = data.replace("tag_name:", "")
                else:
                    locator_details.append(data)

            table = self.driver.find_element(By.XPATH, locator_details[0])

            rows = table.find_elements(By.TAG_NAME, "tr")  # get all of the rows in the table
            cols = len(table.find_elements(By.TAG_NAME, "th")  )# get all of the rows in the table
            rows=rows[1:]
            data=[]
            for row in rows:
                try:
                    # Get the columns (all the column 2)
                    temp=[]
                    for col in range(cols):
                        col = row.find_elements(By.TAG_NAME, "td")[col]  # note: index start from 0, 1 is col 2
                        temp.append(col.text.strip())
                    data.append(temp)
                except:
                    pass


        except Exception as e:
            print(e)
        return data

    def select_backup(self, locator, value):

        dropdown=self.driver.find_element(By.XPATH, locator)
        select = Select(dropdown)
        select.select_by_visible_text(value)
        if (self.log != None):

            self.log.logger.info("Selecting from an element: " + str(locator) + " value selected as : " + str(value))


    def set_backup(self, locator, value):

        if str(locator).endswith("_XPATH"):

            self.driver.find_element(By.XPATH, self.config.get_data_from_a_section("locators", locator)).send_keys(value)

        elif str(locator).endswith("_CSS"):
            self.driver.find_element(By.CSS_SELECTOR, self.config.get_data_from_a_section("locators", locator)).send_keys(
                value)
        elif str(locator).endswith("_ID"):

            self.driver.find_element(By.ID, self.config.get_data_from_a_section("locators", locator)).send_keys(
                value)
        if (self.log != None):

            self.log.logger.info("Typing in an element: " + str(locator) + " value entered as : " + str(value))

    def select_backup2(self, locator, value):
        global dropdown
        if str(locator).endswith("_XPATH"):

            dropdown=self.driver.find_element(By.XPATH, self.config.get_data_from_a_section("locators", locator))
        elif str(locator).endswith("_CSS"):
            dropdown = self.driver.find_element(By.CSS_SELECTOR, self.config.get_data_from_a_section("locators", locator))
        elif str(locator).endswith("_ID"):
            dropdown = self.driver.find_element(By.ID, self.config.get_data_from_a_section("locators", locator))

        select = Select(dropdown)
        select.select_by_visible_text(value)
        if (self.log != None):

            self.log.logger.info("Selecting from an element: " + str(locator) + " value selected as : " + str(value))

    def moveTo(self, locator):

        if str(locator).endswith("_XPATH"):
            element = self.driver.find_element_by_xpath(ConfigReader.readConfig("locators", locator))
        elif str(locator).endswith("_CSS"):
            element = self.driver.find_element_by_css_selector(ConfigReader.readConfig("locators", locator))
        elif str(locator).endswith("_ID"):
            element = self.driver.find_element_by_id(ConfigReader.readConfig("locators", locator))

        action = ActionChains(self.driver)
        action.move_to_element(element).perform()
        if (self.log != None):

            self.log.logger.info("Moving to an element: " + str(locator))
    def enter_details_in_screen(self,screen):
        data=self.excel.get_Alldata_In_Screen(screen)
        columns=data.columns.tolist()
        self.dict_screen_details=self.config_file.get_attributes_of_screen(screen)

        for _, row in data.iterrows():



            for col in columns:
                tag = None
                locator_details = []
                locators = self.dict_screen_details[col]
                for data in locators:
                    if "tag_name:" in data:
                        tag = data.replace("tag_name:", "")
                    else:
                        locator_details.append(data)

                if tag =="input":
                    self.set(locator_details[0], row[col])
                elif tag=="select":
                    self.select(locator_details[0], row[col])

    def func_TCObjective(self,sObjective) :

            if ("Objective" in sObjective):
                sObjective=sObjective.replace("Objective","")

            sObjective = "\nScenario Name: "+ self.config_data.sNewTestCaseName+ "\nDescription: " + sObjective;
            if "excel" in self.log_type.strip().lower() :
                self.func_SetLogMsgWithoutScreenCapture("Objective " + sObjective);

                self.config_data.sTCIDWithObjective[self.config_data.sTestcaseId]=sObjective;
            elif "html" in self.log_type.strip().lower():
                self.func_SetObjective( sObjective);

                self.config_data.sTCIDWithObjective[self.config_data.sTestcaseId]=sObjective;



        
    def enter_details_with_ID_in_screen(self,screen,ID):
        data=self.excel.get_Alldata_With_TCID_In_Screen(screen, ID)
        columns=data.columns.tolist()
        self.dict_screen_details=self.config_file.get_attributes_of_screen(screen)
        columns.remove("TC ID")
        columns.remove("Multiple Run Required")
        self.func_InformationMessageToPrintManualSteps("enter Details in the scrren:" + screen);
        for _, row in data.iterrows():

            for col in columns:
                tag = None
                locator_details = []

                try:
                    locators = self.dict_screen_details[col]

                    for data in locators:
                        if "tag_name:" in data:
                            tag = data.replace("tag_name:", "")
                        else:
                            locator_details.append(data)

                    if tag =="input":
                        self.set(locator_details[0], row[col])
                    elif tag=="select":
                        self.select(locator_details[0], row[col])
                except:
                    pass


    def enterText(self,field):
        title=self.driver.title.strip()

        data=self.excel.get_Alldata_With_TCID_In_Screen(title,self.config_data.sTestcaseId)

        columns=data.columns.tolist()
        self.dict_screen_details=self.config_file.get_attributes_of_screen(title)

        columns.remove("TC ID")
        columns.remove("Multiple Run Required")
        for _, row in data.iterrows():
            for col in columns:
                if field in col:
                    tag = None
                    locator_details = []
                    locators = self.dict_screen_details[col]
                    for data in locators:
                        if "tag_name:" in data:
                            tag = data.replace("tag_name:", "")
                        else:
                            locator_details.append(data)

                    if tag =="input":
                        self.func_InformationMessageToPrintManualSteps(
                            "Enter " + locator_details[0] + " in the text field " + field);
                        self.set(locator_details[0], row[col])

    def clickonField(self,field):

        title = self.driver.title.strip()

        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)
        locator_details = []
        locators = self.dict_screen_details[field]

        tag = None
        for data in locators:
            if "tag_name:" in data:
                tag = data.replace("tag_name:", "")
            else:
                locator_details.append(data)

        self.func_InformationMessageToPrintManualSteps("Click on " + field);
        self.click(locator_details[0])

    @allure.step("Its a test step")
    def step_one(self):

        raise Exception("Step One Pass")

        return False

    @allure.testcase("This is test one")
    @allure.description("This is to reproduce the allure step bug")
    def test_one(self):

        assert 10==20, "It is an error message"
        print("Testcase Failed")
    def SelectText(self, field):

        title = self.driver.title.strip()

        data = self.excel.get_Alldata_With_TCID_In_Screen(title, self.config_data.sTestcaseId)

        columns = data.columns.tolist()
        self.dict_screen_details = self.config_file.get_attributes_of_screen(title)

        columns.remove("TC ID")
        columns.remove("Multiple Run Required")
        for _, row in data.iterrows():
            for col in columns:
                if field in col:
                    tag = None
                    locator_details = []
                    locators = self.dict_screen_details[col]
                    for data in locators:
                        if "tag_name:" in data:
                            tag = data.replace("tag_name:", "")
                        else:
                            locator_details.append(data)


                    if tag == "select":
                        self.func_InformationMessageToPrintManualSteps("Select "+row[col]+" from the dropdown " + field);
                        self.select(locator_details[0], row[col])

    def printPassMsgWscreenshot(self,doc, scriptstatus, msg, screenName):

        new_li = doc.new_tag("li")
        temp_li = doc.new_tag("li")
        temp_li['style'] = 'color: yellowgreen; font-size: 10pt; font-weight: bold;'
        temp_li.string = "PASS " + msg
        scriptstatus.append(temp_li)
        br = doc.new_tag("br")
        scriptstatus.append(br)
        new_i = doc.new_tag("i")
        new_i.string = 'user_screen_snapshot = '
        new_li.append(new_i)
        new_a = doc.new_tag("a")
        new_a['href'] = screenName
        new_img = doc.new_tag("img")
        new_img['height'] = "100"
        new_img['width'] = "160"
        new_img['align'] = "middle"
        new_img.string = 'Click to view full size'
        new_a.append(new_img)
        new_li.append(new_a)
        # new_div.string="New Element"
        scriptstatus.append(new_li)

        br = doc.new_tag("br")
        scriptstatus.append(br)
        return doc

    def printFailMsgWscreenshot(self,doc, scriptstatus, msg, screenName):
        id=self.idcounter
        new_li = doc.new_tag("li")
        temp_li = doc.new_tag("li")
        temp_li['style'] = 'color: red; font-size: 10pt; font-weight: bold;'
        temp_li['id'] = "Screen_" + str(self.idcounter)
        self.li_idcounter.append("Screen_" + str(self.idcounter))
        self.idcounter += 1
        temp_li.string = "FAIL " + msg

        scriptstatus.append(temp_li)
        br = doc.new_tag("br")
        scriptstatus.append(br)
        new_i = doc.new_tag("i")
        new_i.string = 'user_screen_snapshot = '
        new_li.append(new_i)
        new_a = doc.new_tag("a")
        new_a['href'] = screenName
        new_img = doc.new_tag("img")
        new_img['height'] = "100"
        new_img['width'] = "160"
        new_img['align'] = "middle"
        new_img.string = 'Click to view full size'
        new_a.append(new_img)
        new_li.append(new_a)
        # new_div.string="New Element"
        scriptstatus.append(new_li)

        br = doc.new_tag("br")
        scriptstatus.append(br)
        return "Screen_" + str(id)

    def printFailMsgWoscreenshot(self,doc, scriptstatus, msg):

        temp_li = doc.new_tag("li")
        temp_li['style'] = 'color: red; font-size: 10pt; font-weight: bold;'
        temp_li['id'] = "Screen_" + str(self.idcounter)
        self.li_idcounter.append("Screen_" + str(self.idcounter))
        self.idcounter += 1
        temp_li.string = "Step Status: FAIL   Message: " + msg
        scriptstatus.append(temp_li)
        br = doc.new_tag("br")
        scriptstatus.append(br)
        return doc

    def printPassMsgWoscreenshot(self,doc, scriptstatus, msg):
        temp_li = doc.new_tag("li")
        temp_li['style'] = 'color: yellowgreen; font-size: 10pt; font-weight: bold;'
        temp_li.string = "PASS " + msg
        scriptstatus.append(temp_li)
        br = doc.new_tag("br")
        scriptstatus.append(br)
        return doc

    def printInfoMsg(self,doc, scriptstatus, msg):
        temp_li = doc.new_tag("li")
        temp_li['style'] = 'color: blue; font-size: 12pt;'
        temp_li.string = "INFO " + msg
        scriptstatus.append(temp_li)
        br = doc.new_tag("br")
        scriptstatus.append(br)
        return doc
    def printID(self,doc, scriptstatus, msg):
        temp_li = doc.new_tag("li")
        temp_li['style'] = 'color: brown; font-size: 14pt;'
        temp_li.string = msg
        scriptstatus.append(temp_li)
        br = doc.new_tag("br")
        scriptstatus.append(br)
        return doc
    def addTestcaseStatus(self,doc, tcDetails):
        status = doc.find('div', id="status")

        for i, data in enumerate(tcDetails):
            br1 = doc.new_tag("p")
            br1.string = data

            status.append(br1)

    def addTestcasedetails(self,doc, starttime, endtime, logName):
        log = doc.find('table', id="log")
        tr = doc.new_tag("tr")
        td = doc.new_tag("td")
        tr.append(td)
        log.append(tr)
        # ------------------------------
        # tr = doc.new_tag("tr")
        td = doc.new_tag("td")
        td['CLASS'] = "time"
        # td['id']='scripttime'
        td.string = 'Start Time:' + starttime
        tr.append(td)
        td = doc.new_tag("td")
        td['CLASS'] = "note"
        # td['id'] = 'scriptstart'
        td.string = 'Script Start: ' + logName
        tr.append(td)
        log.append(tr)

        # -----------------------------
        tr = doc.new_tag("tr")
        td = doc.new_tag("td")
        td['COLSPAN'] = 3
        ul = doc.new_tag("ul")

        FailDetails=[]
        screenids=[]


        for i, data in enumerate(self.config_data.sLogMsg):

            print("=========screenshotDirectoryFull:",self.config_data.screenshotDirectoryFull,">>",self.config_data.screenshotDirectoryFull+"\\"+self.scriptName+str(self.config_data.iTotalScreenCounter[i])+".png")
            if "PASS" in self.config_data.sTCStatus[i]:
                self.printPassMsgWscreenshot(doc, ul, data, self.config_data.screenshotDirectoryFull+"\\"+self.scriptName+str(self.config_data.iTotalScreenCounter[i])+".png")
            elif "Objective" in self.config_data.sTCStatus[i]:
                self.printID(doc, ul, data)
            elif "INFORMATION" in self.config_data.sTCStatus[i] or "COMMENT" in self.config_data.sTCStatus[i]:
                self.printInfoMsg(doc, ul, data)
            elif "ID" in self.config_data.sTCStatus[i]:
                self.printID(doc, ul, data)
            elif "FAIL" in self.config_data.sTCStatus[i]:
                FailDetails.append(data)
                screenids.append(self.printFailMsgWscreenshot(doc, ul, data, self.config_data.screenshotDirectoryFull+"\\"+self.scriptName+str(self.config_data.iTotalScreenCounter[i])+".png"))

               
        #self.printPassMsgWoscreenshot(doc, ul, "It is a  Pass Message")
        #self.printFailMsgWoscreenshot(doc, ul, "It is a  Fail Message")
        #self.printInfoMsg(doc, ul, "In is an Info Message")
        td.append(ul)
        tr.append(td)
        log.append(tr)
        tr = doc.new_tag("tr")
        td = doc.new_tag("td")
        td['CLASS'] = "pass"
        td.string = "Pass"
        tr.append(td)
        log.append(tr)
        td = doc.new_tag("td")
        td['CLASS'] = "time"
        td.string = "End Time: " + endtime
        tr.append(td)
        log.append(tr)
        td = doc.new_tag("td")
        td['CLASS'] = "note"
        td.string = "Script End: " + logName
        tr.append(td)
        log.append(tr)
        tr = doc.new_tag("tr")
        td = doc.new_tag("td")
        td['COLSPAN'] = 3
        ul = doc.new_tag("ul")
        li = doc.new_tag("li")
        i = doc.new_tag("i")
        i.string = "script_name = " + logName
        li.append(i)
        ul.append(li)
        td.append(ul)
        tr.append(td)
        log.append(tr)
        return FailDetails,screenids

    def addFailureDetails(self,doc,screenids, TestcaseDetails, logName):
        failures = doc.find('div', id="failures")

        br1 = None
        for i, data in enumerate(TestcaseDetails):
            br1 = doc.new_tag("a")
            br1['href'] = "#" + screenids[i]
            br1.string = logName + ": " + data
            br = doc.new_tag("br")
            br1.append(br)

            failures.append(br1)

    def generateHtml(self,doc, filename):
        print("\n===========filename",filename)
        with open(filename, "w") as f:
            f.write(str(doc))

    def readtemplate(self):
        
        with open(os.path.join(self.config_data.projectpath, "Template","log.html"), "r") as f:
            doc = BeautifulSoup(f, "html.parser")
            
        return doc

    def generateHtmlLog(self,opfile,TestcaseDetails,tcWithStatus):

        doc = self.readtemplate()
        dateString = datetime.today().strftime('%d %m %Y')
        logscript = doc.find('td', id="logscript")
        logscript.string = logscript.string + "Test Suit"

        FailDetails,screenids=self.addTestcasedetails(doc, self.config_data.sStartTime, self.func_GetTime(), self.config_data.sNewTestCaseName)

        self.addFailureDetails(doc,screenids, FailDetails, self.config_data.sNewTestCaseName)
        TestcaseDetails =[]
        if len(FailDetails)>0:
            TestcaseDetails.append('Fail: '+self.config_data.sNewTestCaseName)
        else:
            TestcaseDetails.append('Pass: ' + self.config_data.sNewTestCaseName)
        TestcaseDetails.sort()
        try:
            itemlist = []
            if os.path.isfile('outfile'):
                with open('outfile', 'rb') as fp:
                    itemlist = pickle.load(fp)
            itemlist.extend(TestcaseDetails)
            #print("\n=============itemlist",itemlist)
            with open('outfile', 'wb') as fp:
                pickle.dump(itemlist, fp)


        except Exception as e:
            print("Error:", e)
        try:
            self.addTestcaseStatus(doc, TestcaseDetails)


        except Exception as e:
            print("Error:", e)
        

        self.generateHtml(doc, opfile)
        